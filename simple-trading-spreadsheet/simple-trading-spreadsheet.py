"""
This script creates a simple trading spreadsheet in Excel format using the pandas library.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
from datetime import datetime, timedelta
import random

print("Creating DataFrame with specified columns...")
# Create a DataFrame with the specified columns
df = pd.DataFrame(
    columns=[
        "Start Date",  # Start date of the trade
        "End Date",  # End date of the trade
        "Duration",  # Duration of the trade
        "Asset Code",  # Code of the traded asset
        "Strike",  # Strike price of the option (if applicable)
        "Expiration",  # Expiration date of the option (if applicable)
        "Quantity",  # Quantity of the asset traded
        "Initial Unit Price",  # Initial unit price of the trade
        "Final Unit Price",  # Final unit price of the trade
        "Financial Result",  # Financial result of the trade
        "Percentage Result",  # Percentage result of the trade
    ]
)

print("Generating random test data for 5 rows...")
# Generate random test data for 5 rows
rows = []
for _ in range(5):
    # Generate a random start date within the last 100 days
    start_date = datetime.now() - timedelta(days=random.randint(1, 100))
    # Generate an end date that is 1 to 30 days after the start date
    end_date = start_date + timedelta(days=random.randint(1, 30))
    # Generate a random asset code in the format "ASSETXXX"
    asset_code = f"ASSET{random.randint(100, 999)}"
    # Generate a random strike price between 50 and 150
    strike = round(random.uniform(50, 150), 2)
    # Generate an expiration date 1 to 30 days after the end date
    expiration = end_date + timedelta(days=random.randint(1, 30))
    # Generate a random quantity in multiples of 100, between 100 and 1000
    quantity = random.randint(1, 10) * 100
    # Generate a random initial price between 1000 and 5000
    initial_price = round(random.uniform(1, 10), 2)
    # Generate a final price by adding a random variation of -500 to 500 to the initial price
    final_price = round(initial_price + random.uniform(0, 10), 2)

    # Append the generated data to the rows list
    rows.append(
        {
            "Start Date": start_date.strftime("%Y-%m-%d"),
            "End Date": end_date.strftime("%Y-%m-%d"),
            "Duration": "",  # Placeholder for Duration
            "Asset Code": asset_code,
            "Strike": strike,
            "Expiration": expiration.strftime("%Y-%m-%d"),
            "Quantity": quantity,
            "Initial Unit Price": initial_price,
            "Final Unit Price": final_price,
            "Financial Result": "",  # Placeholder for Financial Result
            "Percentage Result": "",  # Placeholder for Percentage Result
        }
    )

# Create the DataFrame from the rows list
df = pd.DataFrame(rows)

print("Defining file path for the Excel file...")
# Define the file path where the Excel file will be saved
file_path = "trade_log.xlsx"  # Save the file in the current working directory

print("Saving DataFrame to Excel file...")
# Save the DataFrame to an Excel file
df.to_excel(file_path, index=False)

print("Loading workbook to add Excel formulas...")
# Load the workbook to add Excel formulas
wb = load_workbook(file_path)
ws = wb.active
if ws is None:
    raise ValueError("No active worksheet found in the workbook.")

# Add formulas for the "Duration" column (column C)
print("Adding Excel formulas for the 'Duration' column...")
for row in range(2, ws.max_row + 1):
    ws[f"C{row}"] = (
        f'=IF(AND(ISNUMBER(DATEVALUE(B{row})), ISNUMBER(DATEVALUE(A{row}))), DATEVALUE(B{row}) - DATEVALUE(A{row}), "")'
    )

# Add formulas for the "Financial Result" column (column J)
print("Adding Excel formulas for the 'Financial Result' column...")
for row in range(2, ws.max_row + 1):
    ws[f"J{row}"] = (
        f'=IF(AND(ISNUMBER(I{row}), ISNUMBER(H{row}), ISNUMBER(G{row})), (I{row}-H{row})*G{row}, "")'
    )

# Add formulas for the "Percentage Result" column (column K)
print("Adding Excel formulas for the 'Percentage Result' column...")
for row in range(2, ws.max_row + 1):
    ws[f"K{row}"] = (
        f'=IF(AND(ISNUMBER(J{row}), ISNUMBER(H{row}), ISNUMBER(G{row})), J{row}/(H{row}*G{row}), "")'
    )

# Format the "Initial Unit Price", "Final Unit Price", "Financial Result", and "Percentage Result" columns
print(
    "Formatting 'Initial Unit Price', 'Final Unit Price', 'Financial Result', and 'Percentage Result' columns..."
)
for row in range(2, ws.max_row + 1):
    ws[f"H{row}"].number_format = (
        numbers.FORMAT_CURRENCY_USD_SIMPLE
    )  # Initial Unit Price
    ws[f"I{row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # Final Unit Price
    ws[f"J{row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # Financial Result
    ws[f"K{row}"].number_format = numbers.FORMAT_PERCENTAGE_00  # Percentage Result

# Save the workbook with the updated formulas and formatting
print("Saving workbook with formulas and formatting...")
wb.save(file_path)

print(f"Spreadsheet saved to: {file_path}")
